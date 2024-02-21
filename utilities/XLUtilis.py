import openpyxl

class GetExcelData:

    def __init__(self,file,sheetname):
        self.file=file
        self.sheetname=sheetname

    def getrows(self):
        workbook=openpyxl.load_workbook(self.file)
        sheet=workbook[self.sheetname]
        return sheet.max_row

    def getcols(self):
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[self.sheetname]
        return sheet.max_column

    def ReadData(self,row,col):
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[self.sheetname]
        return sheet.cell(row,col).value

    def WriteData(self,filename,sheetname,row,col,data):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        sheet.cell(row,col).value = data
        workbook.save(filename=filename)


